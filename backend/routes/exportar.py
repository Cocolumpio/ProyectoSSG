"""Rutas de Exportación de Métricas (Excel + PDF) - DrON Topografía"""
import os
import io
import logging
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.enums import TA_CENTER

from core.config import get_db

db = get_db()
router = APIRouter(prefix="/api")

# --- Exportación de Métricas Históricas ---
import pandas as pd

@router.get("/exportar/metricas-excel")
async def exportar_metricas_excel():
    """
    Exporta las métricas históricas de todos los proyectos a Excel.
    Incluye hojas separadas para: Resumen, Detalle por Proyecto, y Avances Semanales.
    """
    try:
        # Obtener todos los proyectos
        proyectos = await db.proyectos.find({}, {"_id": 0}).to_list(100)
        
        # Crear workbook
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="994B49", end_color="994B49", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # --- Hoja 1: Resumen General ---
        ws_resumen = wb.active
        ws_resumen.title = "Resumen General"
        
        # Headers
        headers_resumen = ["Proyecto", "Ubicación", "Avance %", "Excavación (m³)", "Vol. Planeado (m³)", 
                         "Pilas", "Pilas Plan.", "Anclas", "Anclas Plan.", "Muros", "Muros Plan.", 
                         "Costo Flotilla", "Semanas"]
        
        for col, header in enumerate(headers_resumen, 1):
            cell = ws_resumen.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Datos de proyectos
        total_volumen = 0
        total_pilas = 0
        total_anclas = 0
        total_muros = 0
        total_costo = 0
        
        for row_num, proyecto in enumerate(proyectos, 2):
            proyecto_id = proyecto.get('id')
            
            # Obtener avances
            avances = await db.avances_semanales.find(
                {"proyecto_id": proyecto_id}, {"_id": 0}
            ).to_list(100)
            
            volumen = sum((a.get('volumen_excavacion', 0) or 0) for a in avances)
            pilas = sum((a.get('pilas_completadas', 0) or 0) for a in avances)
            anclas = sum((a.get('anclas_instaladas', 0) or 0) for a in avances)
            muros = sum((a.get('muros_completados', 0) or 0) for a in avances)
            costo = volumen * (proyecto.get('costo_m3', 150) or 150)
            
            total_volumen += volumen
            total_pilas += pilas
            total_anclas += anclas
            total_muros += muros
            total_costo += costo
            
            row_data = [
                proyecto.get('nombre', ''),
                proyecto.get('ubicacion', ''),
                proyecto.get('avance_actual', 0) or 0,
                volumen,
                proyecto.get('volumen_total_planeado', 0) or 0,
                pilas,
                proyecto.get('pilas_planeadas', 0) or 0,
                anclas,
                proyecto.get('anclas_planeadas', 0) or 0,
                muros,
                proyecto.get('muros_planeados', 0) or 0,
                costo,
                len(avances)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws_resumen.cell(row=row_num, column=col, value=value)
                cell.border = border
                if col == 3:  # Avance %
                    cell.number_format = '0.0%'
                    cell.value = value / 100
                elif col in [12]:  # Costo
                    cell.number_format = '$#,##0.00'
        
        # Fila de totales
        total_row = len(proyectos) + 2
        ws_resumen.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
        ws_resumen.cell(row=total_row, column=4, value=total_volumen).font = Font(bold=True)
        ws_resumen.cell(row=total_row, column=6, value=total_pilas).font = Font(bold=True)
        ws_resumen.cell(row=total_row, column=8, value=total_anclas).font = Font(bold=True)
        ws_resumen.cell(row=total_row, column=10, value=total_muros).font = Font(bold=True)
        ws_resumen.cell(row=total_row, column=12, value=total_costo).number_format = '$#,##0.00'
        
        # Ajustar anchos
        for col in ws_resumen.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws_resumen.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)
        
        # --- Hoja 2: Avances Semanales ---
        ws_avances = wb.create_sheet("Avances Semanales")
        
        headers_avances = ["Proyecto", "Semana", "Fecha", "Volumen (m³)", "Pilas", "Anclas", "Muros", "Descripción"]
        for col, header in enumerate(headers_avances, 1):
            cell = ws_avances.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row_num = 2
        for proyecto in proyectos:
            proyecto_id = proyecto.get('id')
            proyecto_nombre = proyecto.get('nombre', '')
            
            avances = await db.avances_semanales.find(
                {"proyecto_id": proyecto_id}, {"_id": 0}
            ).sort("semana", 1).to_list(100)
            
            for avance in avances:
                row_data = [
                    proyecto_nombre,
                    avance.get('semana', 0),
                    avance.get('fecha', ''),
                    avance.get('volumen_excavacion', 0) or 0,
                    avance.get('pilas_completadas', 0) or 0,
                    avance.get('anclas_instaladas', 0) or 0,
                    avance.get('muros_completados', 0) or 0,
                    avance.get('descripcion', '')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws_avances.cell(row=row_num, column=col, value=value)
                    cell.border = border
                
                row_num += 1
        
        # Ajustar anchos
        for col in ws_avances.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws_avances.column_dimensions[col[0].column_letter].width = min(max_length + 2, 25)
        
        # --- Hoja 3: Comparaciones con Residente ---
        ws_comparaciones = wb.create_sheet("Comparaciones Residente")
        
        headers_comp = ["Proyecto", "Fecha", "PDF", "Avance Dron %", "Avance Residente %", 
                       "Discrepancias", "Alerta Enviada"]
        for col, header in enumerate(headers_comp, 1):
            cell = ws_comparaciones.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        comparaciones = await db.comparaciones_avance.find({}, {"_id": 0}).to_list(200)
        
        for row_num, comp in enumerate(comparaciones, 2):
            proyecto = next((p for p in proyectos if p.get('id') == comp.get('proyecto_id')), {})
            row_data = [
                proyecto.get('nombre', 'Desconocido'),
                comp.get('fecha_comparacion', '')[:10] if comp.get('fecha_comparacion') else '',
                comp.get('pdf_nombre', ''),
                comp.get('avance_general_dron', 0),
                comp.get('avance_general_residente', 0),
                len(comp.get('discrepancias_detectadas', [])),
                'Sí' if comp.get('alerta_enviada') else 'No'
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws_comparaciones.cell(row=row_num, column=col, value=value)
                cell.border = border
        
        # Guardar a buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        fecha_actual = datetime.now(timezone.utc).strftime('%Y%m%d')
        filename = f"DrON_Metricas_Historicas_{fecha_actual}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Allow-Origin": "*"
            }
        )
        
    except Exception as e:
        logging.error(f"Error exportando a Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")


@router.get("/exportar/metricas-pdf")
async def exportar_metricas_pdf():
    """
    Exporta las métricas históricas de todos los proyectos a PDF.
    Incluye resumen ejecutivo, gráficos y tablas detalladas.
    """
    try:
        # Obtener todos los proyectos
        proyectos = await db.proyectos.find({}, {"_id": 0}).to_list(100)
        
        # Calcular totales
        total_volumen = 0
        total_pilas = 0
        total_anclas = 0
        total_muros = 0
        total_costo = 0
        proyectos_data = []
        
        for proyecto in proyectos:
            proyecto_id = proyecto.get('id')
            
            avances = await db.avances_semanales.find(
                {"proyecto_id": proyecto_id}, {"_id": 0}
            ).to_list(100)
            
            volumen = sum((a.get('volumen_excavacion', 0) or 0) for a in avances)
            pilas = sum((a.get('pilas_completadas', 0) or 0) for a in avances)
            anclas = sum((a.get('anclas_instaladas', 0) or 0) for a in avances)
            muros = sum((a.get('muros_completados', 0) or 0) for a in avances)
            costo = volumen * (proyecto.get('costo_m3', 150) or 150)
            
            total_volumen += volumen
            total_pilas += pilas
            total_anclas += anclas
            total_muros += muros
            total_costo += costo
            
            proyectos_data.append({
                'nombre': proyecto.get('nombre', ''),
                'ubicacion': proyecto.get('ubicacion', ''),
                'avance': proyecto.get('avance_actual', 0) or 0,
                'volumen': volumen,
                'volumen_plan': proyecto.get('volumen_total_planeado', 0) or 0,
                'pilas': pilas,
                'pilas_plan': proyecto.get('pilas_planeadas', 0) or 0,
                'anclas': anclas,
                'anclas_plan': proyecto.get('anclas_planeadas', 0) or 0,
                'muros': muros,
                'muros_plan': proyecto.get('muros_planeados', 0) or 0,
                'costo': costo,
                'semanas': len(avances)
            })
        
        # Crear PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#994B49'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#994B49'),
            spaceBefore=20,
            spaceAfter=10
        )
        
        story = []
        
        # Título
        story.append(Paragraph("📊 Reporte de Métricas Históricas", title_style))
        story.append(Paragraph(f"DrON Topografía - {datetime.now(timezone.utc).strftime('%d/%m/%Y')}", 
                              ParagraphStyle('Date', parent=styles['Normal'], alignment=TA_CENTER, textColor=colors.gray)))
        story.append(Spacer(1, 30))
        
        # KPIs Resumen
        story.append(Paragraph("Resumen Ejecutivo", subtitle_style))
        
        kpi_data = [
            ["Métrica", "Total"],
            ["Proyectos Activos", str(len(proyectos))],
            ["Excavación Total", f"{total_volumen:,.0f} m³"],
            ["Pilas Completadas", f"{total_pilas:,}"],
            ["Anclas Instaladas", f"{total_anclas:,}"],
            ["Muros Construidos", f"{total_muros:,}"],
            ["Inversión Flotillas", f"${total_costo:,.2f}"]
        ]
        
        kpi_table = Table(kpi_data, colWidths=[200, 150])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#994B49')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FDF2F2')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#994B49')),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 30))
        
        # Tabla de proyectos
        story.append(Paragraph("Detalle por Proyecto", subtitle_style))
        
        proj_headers = ["Proyecto", "Avance", "Excavación", "Pilas", "Anclas", "Muros"]
        proj_data = [proj_headers]
        
        for p in proyectos_data:
            proj_data.append([
                p['nombre'][:20] + '...' if len(p['nombre']) > 20 else p['nombre'],
                f"{p['avance']:.1f}%",
                f"{p['volumen']:,.0f} m³",
                f"{p['pilas']}/{p['pilas_plan']}",
                f"{p['anclas']}/{p['anclas_plan']}",
                f"{p['muros']}/{p['muros_plan']}"
            ])
        
        proj_table = Table(proj_data, colWidths=[120, 60, 80, 60, 60, 60])
        proj_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#994B49')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ]))
        story.append(proj_table)
        story.append(Spacer(1, 30))
        
        # Costos de Flotilla
        story.append(Paragraph("Desglose de Costos de Flotilla", subtitle_style))
        
        costo_headers = ["Proyecto", "Volumen", "Viajes Est.", "Costo Total"]
        costo_data = [costo_headers]
        
        for p in proyectos_data:
            viajes = int(p['volumen'] / 25) if p['volumen'] > 0 else 0
            costo_data.append([
                p['nombre'][:25] + '...' if len(p['nombre']) > 25 else p['nombre'],
                f"{p['volumen']:,.0f} m³",
                f"{viajes:,}",
                f"${p['costo']:,.2f}"
            ])
        
        # Fila de totales
        costo_data.append([
            "TOTAL",
            f"{total_volumen:,.0f} m³",
            "-",
            f"${total_costo:,.2f}"
        ])
        
        costo_table = Table(costo_data, colWidths=[150, 100, 80, 100])
        costo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#994B49')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEE2E2')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        story.append(costo_table)
        
        # Footer
        story.append(Spacer(1, 40))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=TA_CENTER)
        story.append(Paragraph(f"Generado por DrON Topografía - {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC", footer_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        fecha_actual = datetime.now(timezone.utc).strftime('%Y%m%d')
        filename = f"DrON_Metricas_Historicas_{fecha_actual}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Allow-Origin": "*"
            }
        )
        
    except Exception as e:
        logging.error(f"Error exportando a PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")
