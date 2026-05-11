"""Rutas de Catálogo de Maquinaria con IA - DrON Topografía"""
import os
import io
import json
import uuid
import logging
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, UploadFile, File
from openpyxl import load_workbook

from emergentintegrations.llm.chat import LlmChat, UserMessage

from core.config import get_db, EMERGENT_LLM_KEY

db = get_db()
router = APIRouter(prefix="/api")

# --- Catálogo de Maquinaria con IA ---
@router.post("/proyectos/analizar-catalogo-maquinaria")
async def analizar_catalogo_maquinaria(
    file: UploadFile = File(...),
    area_terreno: float = 0,
    volumen_excavacion: float = 0,
    num_pilas: int = 0,
    distancia_pilas: float = 0,
    espacio_maniobra: float = 0
):
    """
    Analiza un catálogo de maquinaria en Excel y usa IA para:
    1. Extraer información de las máquinas
    2. Buscar especificaciones técnicas
    3. Proponer distribución óptima para el proyecto
    """
    import json as json_module
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser Excel (.xlsx o .xls)")
    
    # Leer el archivo Excel
    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content))
    ws = wb.active
    
    # Extraer datos de las máquinas
    maquinas = []
    headers = []
    header_row_found = False
    
    # Mapeo flexible de nombres de columnas
    column_mappings = {
        'tipo': ['TIPO DE MAQUINA', 'TIPO', 'TIPO_MAQUINA', 'MAQUINA', 'EQUIPO', 'TIPO DE EQUIPO'],
        'marca': ['MARCA', 'FABRICANTE', 'MANUFACTURER'],
        'modelo': ['MODELO', 'MODEL', 'NUMERO DE MODELO'],
        'estatus': ['ESTATUS', 'STATUS', 'ESTADO', 'CONDICION', 'FECHA APROX TERMINO PROYECTO A INICIAR'],
        'operador': ['OPERADOR', 'OPERATOR', 'CONDUCTOR'],
        'obra': ['OBRA', 'PROYECTO', 'SITE', 'OBRA ACTUAL'],
        'ubicacion': ['UBICACIÓN', 'UBICACION', 'LOCATION', 'SITIO']
    }
    
    def find_column_value(row_dict, key_names):
        """Busca el valor usando múltiples posibles nombres de columna"""
        for key in key_names:
            if key in row_dict and row_dict[key]:
                return row_dict[key]
        return ''
    
    def is_header_row(row):
        """Detecta si una fila parece ser la fila de headers"""
        row_values = [str(v).strip().upper() if v else '' for v in row]
        # Buscar palabras clave de headers
        keywords = ['TIPO', 'MARCA', 'MODELO', 'MAQUINA', 'EQUIPO']
        matches = sum(1 for kw in keywords for val in row_values if kw in val)
        return matches >= 2  # Al menos 2 coincidencias
    
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        # Saltar filas completamente vacías
        if not any(row):
            continue
        
        # Buscar la fila de headers
        if not header_row_found:
            if is_header_row(row):
                # Normalizar headers: quitar espacios, convertir a mayúsculas
                headers = []
                for i, h in enumerate(row):
                    if h:
                        cleaned = str(h).strip().upper().replace('  ', ' ')
                        headers.append(cleaned)
                    else:
                        headers.append(f"COL_{i}")
                logging.info(f"Headers encontrados en fila {idx}: {headers}")
                header_row_found = True
            continue
        
        # Procesar filas de datos
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = value
        
        # Extraer campos clave usando el mapeo flexible
        tipo = find_column_value(row_dict, column_mappings['tipo'])
        marca = find_column_value(row_dict, column_mappings['marca'])
        modelo = find_column_value(row_dict, column_mappings['modelo'])
        estatus = find_column_value(row_dict, column_mappings['estatus'])
        operador = find_column_value(row_dict, column_mappings['operador'])
        obra = find_column_value(row_dict, column_mappings['obra'])
        ubicacion = find_column_value(row_dict, column_mappings['ubicacion'])
        
        if tipo and str(tipo).strip():
            maquinas.append({
                "tipo": str(tipo).strip(),
                "marca": str(marca).strip() if marca else "",
                "modelo": str(modelo).strip() if modelo else "",
                "estatus": str(estatus).strip() if estatus else "",
                "operador": str(operador).strip() if operador else "",
                "obra_actual": str(obra).strip() if obra else "",
                "ubicacion": str(ubicacion).strip() if ubicacion else ""
            })
    
    logging.info(f"Máquinas encontradas: {len(maquinas)}")
    
    if not maquinas:
        # Dar más información sobre el problema
        raise HTTPException(
            status_code=400, 
            detail=f"No se encontraron máquinas. Headers encontrados: {headers}. Se esperan columnas como: TIPO DE MAQUINA, MARCA, MODELO, ESTATUS"
        )
    
    # Filtrar máquinas disponibles (OPTIMA, SATISFACTORIO, sin estado definido)
    maquinas_disponibles = [
        m for m in maquinas 
        if m.get('estatus', '').upper() not in ['DESHABILITADA', 'EN REPARACION']
    ]
    
    # Categorizar máquinas
    excavadoras = [m for m in maquinas_disponibles if 'EXCAVADORA' in m['tipo'].upper()]
    perforadoras = [m for m in maquinas_disponibles if 'PERFORADORA' in m['tipo'].upper() and 'ANCLA' not in m['tipo'].upper()]
    perforadoras_anclas = [m for m in maquinas_disponibles if 'PERFORADORA ANCLAS' in m['tipo'].upper() or 'ANCLA' in m['tipo'].upper()]
    gruas = [m for m in maquinas_disponibles if 'GRUA' in m['tipo'].upper()]
    manipuladores = [m for m in maquinas_disponibles if 'MANIPULADOR' in m['tipo'].upper()]
    
    # Preparar prompt para IA
    maquinas_texto = ""
    for m in maquinas_disponibles:
        maquinas_texto += f"- {m['tipo']}: {m['marca']} {m['modelo']} (Estado: {m['estatus'] or 'Disponible'})\n"
    
    prompt = f"""Eres un experto en maquinaria de construcción y planificación de obras.

CATÁLOGO DE MAQUINARIA DISPONIBLE:
{maquinas_texto}

DATOS DEL PROYECTO:
- Área del terreno: {area_terreno} m²
- Volumen de excavación: {volumen_excavacion} m³
- Número de pilas a perforar: {num_pilas}
- Distancia entre pilas: {distancia_pilas} m
- Espacio de maniobra disponible: {espacio_maniobra} m²

TAREA:
1. Para cada máquina del catálogo, proporciona las especificaciones técnicas aproximadas:
   - Dimensiones (largo x ancho x altura en metros)
   - Radio de giro
   - Rendimiento estimado (m³/hora para excavadoras, pilas/día para perforadoras)
   - Peso operativo

2. Analiza qué máquinas son más adecuadas para este proyecto considerando:
   - Espacio disponible para maniobras
   - Volumen de trabajo
   - Eficiencia esperada

3. Propón un PLAN DE EJECUCIÓN ÓPTIMO:
   - FASE 1 EXCAVACIÓN: Qué excavadoras usar y en qué orden
   - FASE 2 PERFORACIÓN DE PILAS: Qué perforadoras usar
   - FASE 3 ANCLAS: Qué perforadoras de anclas usar
   - Distribución espacial recomendada

4. Calcula:
   - Tiempo estimado para excavación
   - Tiempo estimado para perforación de pilas
   - Tiempo estimado para anclas
   - Recomendaciones para optimizar el uso del espacio

Responde en formato JSON con esta estructura:
{{
    "maquinas_con_specs": [
        {{
            "tipo": "...",
            "marca": "...",
            "modelo": "...",
            "dimensiones": {{"largo": 0, "ancho": 0, "altura": 0}},
            "radio_giro": 0,
            "rendimiento": "...",
            "peso_operativo": 0,
            "adecuada_para_proyecto": true/false,
            "razon": "..."
        }}
    ],
    "plan_excavacion": {{
        "maquinas_recomendadas": ["..."],
        "estrategia": "...",
        "tiempo_estimado_dias": 0,
        "rendimiento_esperado_m3_dia": 0
    }},
    "plan_pilas": {{
        "maquinas_recomendadas": ["..."],
        "estrategia": "...",
        "tiempo_estimado_dias": 0,
        "pilas_por_dia": 0
    }},
    "plan_anclas": {{
        "maquinas_recomendadas": ["..."],
        "estrategia": "...",
        "tiempo_estimado_dias": 0,
        "anclas_por_dia": 0
    }},
    "distribucion_espacial": {{
        "recomendacion": "...",
        "zonas_trabajo": ["..."],
        "consideraciones_seguridad": ["..."]
    }},
    "resumen_ejecutivo": "..."
}}
"""
    
    try:
        from emergentintegrations.llm.chat import UserMessage
        
        # Crear chat con Gemini
        llm = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"catalogo-maquinaria-{uuid.uuid4()}",
            system_message="Eres un experto en maquinaria de construcción y planificación de obras. Siempre respondes en formato JSON válido."
        ).with_model("gemini", "gemini-2.0-flash")
        
        # Enviar mensaje
        user_message = UserMessage(text=prompt)
        response = await llm.send_message(user_message)
        
        logging.info(f"Respuesta IA recibida: {len(response)} caracteres")
        
        # Parsear respuesta JSON
        response_text = response.strip()
        if response_text.startswith("```json"):
            response_text = response_text[7:]
        if response_text.startswith("```"):
            response_text = response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        
        # Intentar extraer JSON de la respuesta
        import re
        json_match = re.search(r'\{[\s\S]*\}', response_text)
        if json_match:
            resultado_ia = json_module.loads(json_match.group())
        else:
            resultado_ia = json_module.loads(response_text.strip())
        
        return {
            "success": True,
            "total_maquinas": len(maquinas),
            "maquinas_disponibles": len(maquinas_disponibles),
            "resumen_catalogo": {
                "excavadoras": len(excavadoras),
                "perforadoras": len(perforadoras),
                "perforadoras_anclas": len(perforadoras_anclas),
                "gruas": len(gruas),
                "manipuladores": len(manipuladores)
            },
            "maquinas_raw": maquinas_disponibles,
            "analisis_ia": resultado_ia
        }
    except json_module.JSONDecodeError as e:
        logging.warning(f"Error parseando JSON de IA: {e}")
        # Si no puede parsear JSON, devolver texto plano
        return {
            "success": True,
            "total_maquinas": len(maquinas),
            "maquinas_disponibles": len(maquinas_disponibles),
            "resumen_catalogo": {
                "excavadoras": len(excavadoras),
                "perforadoras": len(perforadoras),
                "perforadoras_anclas": len(perforadoras_anclas),
                "gruas": len(gruas),
                "manipuladores": len(manipuladores)
            },
            "maquinas_raw": maquinas_disponibles,
            "analisis_ia_texto": response if 'response' in locals() else "Error procesando respuesta de IA",
            "mensaje": "Catálogo procesado. El análisis de IA está disponible como texto."
        }
    except Exception as e:
        logging.error(f"Error analizando catálogo con IA: {e}")
        import traceback
        traceback.print_exc()
        # Devolver éxito parcial - tenemos las máquinas pero no el análisis IA
        return {
            "success": True,
            "error_ia": str(e),
            "total_maquinas": len(maquinas),
            "maquinas_disponibles": len(maquinas_disponibles),
            "maquinas_raw": maquinas_disponibles,
            "resumen_catalogo": {
                "excavadoras": len(excavadoras),
                "perforadoras": len(perforadoras),
                "perforadoras_anclas": len(perforadoras_anclas),
                "gruas": len(gruas),
                "manipuladores": len(manipuladores)
            },
            "mensaje": f"Catálogo procesado exitosamente ({len(maquinas_disponibles)} máquinas disponibles). El análisis con IA no está disponible temporalmente."
        }


@router.post("/proyectos/{proyecto_id}/guardar-catalogo-maquinaria")
async def guardar_catalogo_maquinaria(proyecto_id: str, data: dict):
    """
    Guarda el catálogo de maquinaria y el análisis de IA en el proyecto.
    """
    proyecto = await db.proyectos.find_one({"id": proyecto_id})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    await db.proyectos.update_one(
        {"id": proyecto_id},
        {"$set": {
            "catalogo_maquinaria": data.get("maquinas", []),
            "analisis_maquinaria_ia": data.get("analisis_ia", {}),
            "parametros_proyecto": data.get("parametros", {}),
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"success": True, "message": "Catálogo guardado correctamente"}


@router.get("/proyectos/{proyecto_id}/catalogo-maquinaria")
async def obtener_catalogo_maquinaria(proyecto_id: str):
    """Obtiene el catálogo de maquinaria de un proyecto"""
    proyecto = await db.proyectos.find_one(
        {"id": proyecto_id}, 
        {"_id": 0, "catalogo_maquinaria": 1, "analisis_maquinaria_ia": 1, "parametros_proyecto": 1}
    )
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    return {
        "catalogo": proyecto.get("catalogo_maquinaria", []),
        "analisis_ia": proyecto.get("analisis_maquinaria_ia", {}),
        "parametros": proyecto.get("parametros_proyecto", {})
    }


# --- Comparación de Plan IA vs Cronograma del Usuario ---
@router.post("/proyectos/{proyecto_id}/comparar-plan-ia")
async def comparar_plan_ia_vs_cronograma(proyecto_id: str):
    """
    Compara el plan generado por IA vs el cronograma planificado por el usuario.
    Usa IA para analizar y determinar si el plan propuesto es mejor, igual o peor.
    """
    import json as json_module
    
    proyecto = await db.proyectos.find_one({"id": proyecto_id}, {"_id": 0})
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    analisis_ia = proyecto.get("analisis_maquinaria_ia", {})
    if not analisis_ia:
        raise HTTPException(status_code=400, detail="No hay análisis de IA para este proyecto. Primero sube el catálogo de maquinaria.")
    
    # Obtener datos del cronograma del usuario
    semanas_excavacion = proyecto.get("semanas_excavacion", 0)
    semanas_pilas = proyecto.get("semanas_pilas", 0)
    semanas_muros = proyecto.get("semanas_muros", 0)
    semanas_planeadas = proyecto.get("semanas_planeadas", 0)
    volumen_total = proyecto.get("volumen_total_planeado", 0)
    pilas_planeadas = proyecto.get("pilas_planeadas", 0)
    anclas_planeadas = proyecto.get("anclas_planeadas", 0)
    
    # Obtener datos del avance real
    avances = await db.avances_semanales.find({"proyecto_id": proyecto_id}, {"_id": 0}).to_list(100)
    semanas_reales = len(avances)
    volumen_real = sum((a.get("volumen_excavacion", 0) or 0) for a in avances)
    pilas_real = sum((a.get("pilas_completadas", 0) or 0) for a in avances)
    anclas_real = sum((a.get("anclas_instaladas", 0) or 0) for a in avances)
    
    # Datos del plan de IA
    plan_excavacion_ia = analisis_ia.get("plan_excavacion", {})
    plan_pilas_ia = analisis_ia.get("plan_pilas", {})
    plan_anclas_ia = analisis_ia.get("plan_anclas", {})
    
    dias_excavacion_ia = plan_excavacion_ia.get("tiempo_estimado_dias", 0)
    dias_pilas_ia = plan_pilas_ia.get("tiempo_estimado_dias", 0)
    dias_anclas_ia = plan_anclas_ia.get("tiempo_estimado_dias", 0)
    
    # Convertir días a semanas (5 días laborales = 1 semana)
    semanas_excavacion_ia = round(dias_excavacion_ia / 5, 1) if dias_excavacion_ia else 0
    semanas_pilas_ia = round(dias_pilas_ia / 5, 1) if dias_pilas_ia else 0
    semanas_anclas_ia = round(dias_anclas_ia / 5, 1) if dias_anclas_ia else 0
    
    # Crear prompt para análisis comparativo
    prompt = f"""Eres un experto en planificación de obras de construcción.

DATOS DEL PROYECTO: {proyecto.get('nombre', 'Sin nombre')}

CRONOGRAMA PLANIFICADO POR EL USUARIO:
- Semanas para excavación: {semanas_excavacion} semanas
- Semanas para pilas: {semanas_pilas} semanas
- Semanas para anclas/muros: {semanas_muros} semanas
- Total semanas planeadas: {semanas_planeadas} semanas
- Volumen a excavar: {volumen_total} m³
- Pilas a perforar: {pilas_planeadas}
- Anclas a instalar: {anclas_planeadas}

PLAN GENERADO POR IA:
- Semanas para excavación: {semanas_excavacion_ia} semanas ({dias_excavacion_ia} días)
- Semanas para pilas: {semanas_pilas_ia} semanas ({dias_pilas_ia} días)  
- Semanas para anclas: {semanas_anclas_ia} semanas ({dias_anclas_ia} días)
- Rendimiento excavación: {plan_excavacion_ia.get('rendimiento_esperado_m3_dia', 0)} m³/día
- Rendimiento pilas: {plan_pilas_ia.get('pilas_por_dia', 0)} pilas/día
- Rendimiento anclas: {plan_anclas_ia.get('anclas_por_dia', 0)} anclas/día
- Máquinas recomendadas excavación: {plan_excavacion_ia.get('maquinas_recomendadas', [])}
- Máquinas recomendadas pilas: {plan_pilas_ia.get('maquinas_recomendadas', [])}

AVANCE REAL HASTA AHORA (Semana {semanas_reales}):
- Volumen excavado: {volumen_real} m³
- Pilas completadas: {pilas_real}
- Anclas instaladas: {anclas_real}

TAREA:
1. Compara los tres escenarios: Plan del Usuario, Plan IA, y Avance Real
2. Determina cuál plan es más realista y eficiente
3. Analiza si el avance real va acorde a alguno de los planes
4. Proporciona recomendaciones para optimizar

Responde en formato JSON:
{{
    "comparacion_general": {{
        "plan_usuario_dias_total": <número>,
        "plan_ia_dias_total": <número>,
        "diferencia_dias": <número positivo si IA es más rápido>,
        "porcentaje_mejora": <porcentaje de mejora del plan IA vs usuario>
    }},
    "evaluacion_excavacion": {{
        "usuario_semanas": {semanas_excavacion},
        "ia_semanas": {semanas_excavacion_ia},
        "mejor_plan": "<usuario | ia | similar>",
        "razon": "<explicación>"
    }},
    "evaluacion_pilas": {{
        "usuario_semanas": {semanas_pilas},
        "ia_semanas": {semanas_pilas_ia},
        "mejor_plan": "<usuario | ia | similar>",
        "razon": "<explicación>"
    }},
    "evaluacion_anclas": {{
        "usuario_semanas": {semanas_muros},
        "ia_semanas": {semanas_anclas_ia},
        "mejor_plan": "<usuario | ia | similar>",
        "razon": "<explicación>"
    }},
    "estado_avance_real": {{
        "porcentaje_excavacion": <% completado>,
        "porcentaje_pilas": <% completado>,
        "porcentaje_anclas": <% completado>,
        "alineado_con": "<plan_usuario | plan_ia | retrasado | adelantado>",
        "semanas_restantes_estimadas": <número>
    }},
    "veredicto": "<PLAN_IA_MEJOR | PLAN_USUARIO_MEJOR | SIMILAR>",
    "confianza": "<ALTA | MEDIA | BAJA>",
    "resumen": "<resumen ejecutivo de 2-3 oraciones>",
    "recomendaciones": ["<recomendación 1>", "<recomendación 2>", ...]
}}"""
    
    try:
        llm = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"comparacion-plan-{proyecto_id}-{uuid.uuid4()}",
            system_message="Eres un experto en planificación de obras de construcción. Siempre respondes en JSON válido."
        ).with_model("gemini", "gemini-2.0-flash")
        
        user_message = UserMessage(text=prompt)
        response = await llm.send_message(user_message)
        
        # Parsear respuesta
        response_text = response.strip()
        if response_text.startswith("```json"):
            response_text = response_text[7:]
        if response_text.startswith("```"):
            response_text = response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        
        import re
        json_match = re.search(r'\{[\s\S]*\}', response_text)
        if json_match:
            comparacion_ia = json_module.loads(json_match.group())
        else:
            comparacion_ia = json_module.loads(response_text.strip())
        
        # Guardar comparación en el proyecto
        comparacion_completa = {
            "fecha_comparacion": datetime.now(timezone.utc).isoformat(),
            "datos_usuario": {
                "semanas_excavacion": semanas_excavacion,
                "semanas_pilas": semanas_pilas,
                "semanas_anclas": semanas_muros,
                "semanas_total": semanas_planeadas,
                "volumen_total": volumen_total,
                "pilas_planeadas": pilas_planeadas,
                "anclas_planeadas": anclas_planeadas
            },
            "datos_ia": {
                "semanas_excavacion": semanas_excavacion_ia,
                "semanas_pilas": semanas_pilas_ia,
                "semanas_anclas": semanas_anclas_ia,
                "semanas_total": semanas_excavacion_ia + semanas_pilas_ia + semanas_anclas_ia,
                "rendimiento_excavacion_m3_dia": plan_excavacion_ia.get("rendimiento_esperado_m3_dia", 0),
                "rendimiento_pilas_dia": plan_pilas_ia.get("pilas_por_dia", 0),
                "rendimiento_anclas_dia": plan_anclas_ia.get("anclas_por_dia", 0)
            },
            "datos_reales": {
                "semanas_transcurridas": semanas_reales,
                "volumen_excavado": volumen_real,
                "pilas_completadas": pilas_real,
                "anclas_instaladas": anclas_real
            },
            "analisis_ia": comparacion_ia
        }
        
        await db.proyectos.update_one(
            {"id": proyecto_id},
            {"$set": {
                "comparacion_planes": comparacion_completa,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        return {
            "success": True,
            "comparacion": comparacion_completa
        }
        
    except Exception as e:
        logging.error(f"Error comparando planes: {e}")
        import traceback
        traceback.print_exc()
        
        # Devolver comparación básica sin IA
        comparacion_basica = {
            "fecha_comparacion": datetime.now(timezone.utc).isoformat(),
            "datos_usuario": {
                "semanas_excavacion": semanas_excavacion,
                "semanas_pilas": semanas_pilas,
                "semanas_anclas": semanas_muros,
                "semanas_total": semanas_planeadas
            },
            "datos_ia": {
                "semanas_excavacion": semanas_excavacion_ia,
                "semanas_pilas": semanas_pilas_ia,
                "semanas_anclas": semanas_anclas_ia,
                "semanas_total": semanas_excavacion_ia + semanas_pilas_ia + semanas_anclas_ia
            },
            "datos_reales": {
                "semanas_transcurridas": semanas_reales,
                "volumen_excavado": volumen_real,
                "pilas_completadas": pilas_real,
                "anclas_instaladas": anclas_real
            },
            "error_ia": str(e)
        }
        
        await db.proyectos.update_one(
            {"id": proyecto_id},
            {"$set": {
                "comparacion_planes": comparacion_basica,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        return {
            "success": True,
            "comparacion": comparacion_basica,
            "mensaje": "Comparación guardada. El análisis de IA no está disponible."
        }


@router.get("/proyectos/{proyecto_id}/comparacion-planes")
async def obtener_comparacion_planes(proyecto_id: str):
    """Obtiene la comparación de planes guardada para un proyecto"""
    proyecto = await db.proyectos.find_one(
        {"id": proyecto_id},
        {"_id": 0, "comparacion_planes": 1, "nombre": 1}
    )
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    return {
        "proyecto_nombre": proyecto.get("nombre", ""),
        "comparacion": proyecto.get("comparacion_planes", None)
    }


@router.get("/dashboard/comparaciones-resumen")
async def obtener_comparaciones_dashboard():
    """Obtiene un resumen de comparaciones de todos los proyectos para el dashboard"""
    proyectos = await db.proyectos.find(
        {"comparacion_planes": {"$exists": True}},
        {"_id": 0, "id": 1, "nombre": 1, "comparacion_planes": 1, "avance_actual": 1}
    ).to_list(100)
    
    resumen = []
    for p in proyectos:
        comp = p.get("comparacion_planes", {})
        if comp:
            resumen.append({
                "proyecto_id": p["id"],
                "proyecto_nombre": p.get("nombre", ""),
                "avance_actual": p.get("avance_actual", 0),
                "datos_usuario": comp.get("datos_usuario", {}),
                "datos_ia": comp.get("datos_ia", {}),
                "datos_reales": comp.get("datos_reales", {}),
                "veredicto": comp.get("analisis_ia", {}).get("veredicto", "NO_ANALIZADO"),
                "fecha_comparacion": comp.get("fecha_comparacion", "")
            })
    
    return {"proyectos": resumen}

