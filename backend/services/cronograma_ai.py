"""
Servicio de análisis de cronogramas con IA y detección de avance en fotos
"""
import os
import logging
import base64
import pandas as pd
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Any, Optional
from io import BytesIO

from dotenv import load_dotenv
load_dotenv()

EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY")


def excel_date_to_string(excel_date: int) -> str:
    """Convierte fecha serial de Excel a string YYYY-MM-DD"""
    try:
        if isinstance(excel_date, (int, float)) and excel_date > 40000:
            # Excel serial date (days since 1899-12-30)
            base_date = datetime(1899, 12, 30)
            result_date = base_date + timedelta(days=int(excel_date))
            return result_date.strftime("%Y-%m-%d")
        return str(excel_date)
    except Exception:
        return str(excel_date)


# --- Mapeo de categorías de Programa de Obra (formato V2) a fases DrON ---
# Cada categoría se mapea a (fase_dron, unidad_esperada) o None si es general/no-físico
_CATEGORIA_PROGRAMA_MAPPING = {
    "EXCAVACION": "excavacion",
    "EXCAVACIÓN": "excavacion",
    "EXCAVACION Y CARGA DE MATERIAL": "excavacion",
    "EXCAVACIÓN Y CARGA DE MATERIAL": "excavacion",
    "PILAS DE CIMENTACION": "pilas",
    "PILAS DE CIMENTACIÓN": "pilas",
    "PILAS": "pilas",
    "PERFILES DE CIMENTACION": "pilas",  # también pilas (perfiles estructurales)
    "PERFILES DE CIMENTACIÓN": "pilas",
    "REFORZAMIENTO DE COLINDANCIAS": "pilas",  # Clemente: pilas de contención
    "ANCLAJE": "anclas",
    "ANCLAS": "anclas",
    "ANCLAJES": "anclas",
    "ESTABILIZACION": "anclas",  # Clemente: aquí van los anclajes
    "ESTABILIZACIÓN": "anclas",  # OJO: en Prados V2 esto era "muros" - aquí prevalece "anclas"
    "MUROS": "muros",
    "MURO": "muros",
    "MUROS DE CONTENCION": "muros",
    "MUROS DE CONTENCIÓN": "muros",
    "MURO DE CONTENCION": "muros",
    "MURO DE CONTENCIÓN": "muros",
    "ZAPATA CORRIDA": "cimentacion",
    "ZAPATA LINDERO": "cimentacion",
    "CIMENTACION": "cimentacion",
    "CIMENTACIÓN": "cimentacion",
    "GENERALES": "generales",
    "PRELIMINARES": "generales",
}


def _normalizar_categoria(nombre: str) -> str:
    """Normaliza el nombre de la categoría para hacer matching robusto."""
    if not nombre:
        return ""
    return re.sub(r"\s+", " ", str(nombre)).strip().upper()


def parse_excel_programa_obra(file_content: bytes) -> Optional[Dict[str, Any]]:
    """
    Parser para programas de obra estilo "Avance por Actividades" (formato V2).
    Detecta automáticamente la presencia del formato y devuelve None si no aplica.

    Estructura esperada:
      - Fechas de inicio / fin / semanas en filas 8-10
      - Fila con headers ``CONCEPTO | UNIDAD | CANTIDAD | IMPORTE DE CONTRATO``
        (suele estar en torno a la fila 20).
      - Cada categoría (GENERALES, EXCAVACION, ANCLAJE, MUROS, PILAS, ...)
        en columna C con ``importe`` en col F y sin ``unidad``.
      - Cada partida bajo la categoría tiene ``unidad`` (M3, M2, PZA, PZAS, ML, ...)
        y ``cantidad`` numérica.
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    except Exception:
        return None

    # Buscar la primera hoja con headers válidos
    ws = None
    header_row = None
    for sheet in wb.sheetnames:
        candidate = wb[sheet]
        for r in range(1, min(80, candidate.max_row + 1)):
            row_vals = [candidate.cell(row=r, column=c).value for c in range(1, 12)]
            row_str = " | ".join(str(v).upper() for v in row_vals if v is not None)
            # Acepta header con CONCEPTO + UNIDAD + (CANTIDAD o VOLUMEN o PRESUPUESTO).
            # El campo IMPORTE es opcional — algunos formatos solo tienen volumen/presupuesto.
            tiene_concepto = "CONCEPTO" in row_str
            tiene_unidad = "UNIDAD" in row_str
            tiene_cantidad = (
                "CANTIDAD" in row_str
                or "VOLUMEN DE PRESUPUESTO" in row_str
                or "VOLUMEN PRESUPUESTO" in row_str
                or "PRESUPUESTO" in row_str  # Torre Mezquitan format
            )
            if tiene_concepto and tiene_unidad and tiene_cantidad:
                ws = candidate
                header_row = r
                break
        if ws:
            break

    if not ws or not header_row:
        return None

    # Determinar columnas de concepto / unidad / cantidad / importe
    col_concepto = col_unidad = col_cantidad = col_importe = None
    for c in range(1, 15):
        val = ws.cell(row=header_row, column=c).value
        if not val:
            continue
        s = str(val).strip().upper()
        if s == "CONCEPTO":
            col_concepto = c
        elif s == "UNIDAD":
            col_unidad = c
        elif s in ("CANTIDAD", "VOLUMEN DE PRESUPUESTO", "VOLUMEN PRESUPUESTO", "PRESUPUESTO"):
            col_cantidad = c
        elif "IMPORTE" in s and col_importe is None:
            col_importe = c

    if not (col_concepto and col_unidad and col_cantidad):
        return None

    # Extraer metadatos del proyecto (fechas + semanas) buscando keywords en cols B-D
    fecha_inicio = None
    fecha_fin = None
    semanas_planeadas = 0
    for r in range(1, header_row):
        for c in range(1, 8):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            label = v.upper()
            if "INICIO DE CONTRATO" in label or "INICIO REAL" in label:
                for cc in range(c + 1, min(c + 4, 12)):
                    cand = ws.cell(row=r, column=cc).value
                    if isinstance(cand, datetime):
                        if not fecha_inicio:
                            fecha_inicio = cand.strftime("%Y-%m-%d")
                        break
            elif "TERMINO DE CONTRATO" in label or "TÉRMINO DE CONTRATO" in label:
                for cc in range(c + 1, min(c + 4, 12)):
                    cand = ws.cell(row=r, column=cc).value
                    if isinstance(cand, datetime):
                        fecha_fin = cand.strftime("%Y-%m-%d")
                        break
            elif "SEMANAS DE EJECUCIÓN" in label or "SEMANAS DE EJECUCION" in label:
                # Solo tomar el primer match (suele ser el total de semanas del proyecto)
                if semanas_planeadas:
                    continue
                for cc in range(c + 1, min(c + 4, 12)):
                    cand = ws.cell(row=r, column=cc).value
                    if isinstance(cand, (int, float)) and 0 < cand < 500:
                        semanas_planeadas = int(cand)
                        break

    # ---- Detectar rangos de columnas por semana ----
    # Soporta dos variantes:
    #  • Prados V2: "PRELIMINARES (SEMANA 0)" + número plano + 1 col separadora (8 cols/semana)
    #  • Clemente:  "SEMANA 01", "SEMANA 02"... consecutivas (7 cols/semana, sin separadora)
    semanas_cols = []
    # Identificar primero la fila de etiquetas de día y la de fechas (suelen ser header_row+1 y header_row+2)
    day_label_row = None
    fecha_row = None
    for r in range(header_row + 1, header_row + 5):
        for c in range(1, min(40, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().upper() in ("LUN", "LUNES") and day_label_row is None:
                day_label_row = r
            if isinstance(v, datetime) and fecha_row is None:
                fecha_row = r
        if day_label_row and fecha_row:
            break

    import re as _re
    semana_re = _re.compile(r"^\s*SEMANA\s*0*(\d{1,3})\s*$", _re.IGNORECASE)

    # Encontrar primera columna válida (LUN + fecha)
    primera_col_lun = None
    if day_label_row and fecha_row:
        for c in range(1, ws.max_column + 1):
            lbl = ws.cell(row=day_label_row, column=c).value
            fd = ws.cell(row=fecha_row, column=c).value
            if isinstance(lbl, str) and lbl.strip().upper() in ("LUN", "LUNES") and isinstance(fd, datetime):
                primera_col_lun = c
                break

    for c in range(max(primera_col_lun or 1, 1), ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        # Esta columna debe ser inicio de semana (LUN + fecha)
        es_lun = False
        es_fecha = False
        if day_label_row:
            lbl = ws.cell(row=day_label_row, column=c).value
            es_lun = isinstance(lbl, str) and lbl.strip().upper() in ("LUN", "LUNES")
        if fecha_row:
            fd = ws.cell(row=fecha_row, column=c).value
            es_fecha = isinstance(fd, datetime)
        if not (es_lun and es_fecha):
            continue

        # Determinar el número de semana
        semana_num = None
        if isinstance(v, str):
            v_up = v.upper()
            if "PRELIMINARES" in v_up:
                semana_num = 0
            else:
                m = semana_re.match(v.strip())
                if m:
                    semana_num = int(m.group(1))
        elif isinstance(v, (int, float)) and 0 < v <= 100 and v == int(v):
            semana_num = int(v)

        if semana_num is None:
            continue
        semanas_cols.append({"semana": semana_num, "col_inicio": c, "col_fin": c + 6})

    # Capturar fechas inicio/fin de cada semana

    for s in semanas_cols:
        f_i = ws.cell(row=fecha_row, column=s["col_inicio"]).value if fecha_row else None
        f_f = ws.cell(row=fecha_row, column=s["col_fin"]).value if fecha_row else None
        s["fecha_inicio"] = f_i.strftime("%Y-%m-%d") if isinstance(f_i, datetime) else None
        s["fecha_fin"] = f_f.strftime("%Y-%m-%d") if isinstance(f_f, datetime) else s["fecha_inicio"]

    # Recorrer las filas de datos
    categorias = {}
    categoria_actual = None
    presupuesto_total = 0.0
    nombre_proyecto = None

    # Intentar obtener el nombre del proyecto (suele estar en C3 o cerca)
    for r in range(1, header_row):
        v = ws.cell(row=r, column=3).value
        if isinstance(v, str) and v.strip() and "ESPACIO" not in v.upper() and "TIEMPOS" not in v.upper() and "AVANCE" not in v.upper() and "FECHA" not in v.upper() and "SEMANAS" not in v.upper():
            nombre_proyecto = v.strip()
            break
        # Fallback al nombre "Espacio prados monraz" (col C3)
        if isinstance(v, str) and "ESPACIO" in v.upper() and not nombre_proyecto:
            nombre_proyecto = v.strip()

    for r in range(header_row + 1, ws.max_row + 1):
        concepto = ws.cell(row=r, column=col_concepto).value
        unidad = ws.cell(row=r, column=col_unidad).value
        cantidad = ws.cell(row=r, column=col_cantidad).value
        importe = ws.cell(row=r, column=col_importe).value if col_importe else None

        if not isinstance(concepto, str) or not concepto.strip():
            continue
        concepto_clean = concepto.strip()
        if concepto_clean.upper().startswith("TOTALES"):
            if isinstance(importe, (int, float)):
                presupuesto_total = float(importe)
            break

        cat_norm = _normalizar_categoria(concepto_clean)
        # Heurística: una categoría es una fila SIN unidad y SIN cantidad numérica,
        # con texto corto (suelen ser headers de sección como EXCAVACION, ANCLAJE, MUROS DE CONTENCION).
        # Acepta tanto la variante con importe (Prados V2) como sin importe (Clemente).
        sin_unidad = unidad is None or str(unidad).strip() == ""
        sin_cantidad = not isinstance(cantidad, (int, float)) or cantidad == 0
        es_categoria = (
            sin_unidad
            and sin_cantidad
            and "\n" not in concepto_clean
            and len(concepto_clean) < 60
        )

        if es_categoria:
            categoria_actual = concepto_clean
            fase = _CATEGORIA_PROGRAMA_MAPPING.get(cat_norm, "otros")
            categorias[categoria_actual] = {
                "nombre": categoria_actual,
                "fase": fase,
                "importe": float(importe or 0) if isinstance(importe, (int, float)) else 0.0,
                "items": [],
            }
            continue

        # Es una partida
        if categoria_actual and isinstance(cantidad, (int, float)) and cantidad > 0:
            unidad_str = str(unidad).strip() if unidad else ""

            # Capturar el desglose por semana (sumando las celdas diarias del item)
            semanas_planeadas_item = []
            for s in semanas_cols:
                suma_semana = 0.0
                for col in range(s["col_inicio"], s["col_fin"] + 1):
                    val = ws.cell(row=r, column=col).value
                    if isinstance(val, (int, float)):
                        suma_semana += float(val)
                semanas_planeadas_item.append({
                    "semana": s["semana"],
                    "cantidad": round(suma_semana, 4),
                })

            categorias[categoria_actual]["items"].append({
                "descripcion": concepto_clean,
                "unidad": unidad_str,
                "cantidad": float(cantidad),
                "importe": float(importe) if isinstance(importe, (int, float)) else 0.0,
                "semanas": semanas_planeadas_item,
            })

    # Post-procesamiento: refinar la fase de cada categoría según la unidad
    # mayoritaria de sus partidas. Esto resuelve casos ambiguos (p. ej.
    # "ESTABILIZACION" puede ser muros M2 en un proyecto o anclajes PZA en otro).
    for cat_data in categorias.values():
        # Si los items son todos M3 → excavación; todos M2 → muros; todos PZA → anclas/pilas
        unidades = [it["unidad"].upper() for it in cat_data["items"]]
        if not unidades:
            continue
        cat_norm_local = _normalizar_categoria(cat_data["nombre"])
        # Solo refinar si la categoría no tiene mapping explícito fuerte
        if cat_data["fase"] == "otros" or cat_norm_local in (
            "ESTABILIZACION", "ESTABILIZACIÓN", "REFORZAMIENTO DE COLINDANCIAS"
        ):
            unidad_mayor = max(set(unidades), key=unidades.count)
            if unidad_mayor in ("M3", "M³"):
                cat_data["fase"] = "excavacion"
            elif unidad_mayor in ("M2", "M²"):
                cat_data["fase"] = "muros"
            elif unidad_mayor in ("PZA", "PZAS", "PIEZA", "PIEZAS"):
                # Heurística: si el nombre menciona "pila" o "colindancia", es pila;
                # de lo contrario, es ancla.
                nombre_up = cat_data["nombre"].upper()
                if "PILA" in nombre_up or "COLINDANCIA" in nombre_up or "PERFIL" in nombre_up:
                    cat_data["fase"] = "pilas"
                else:
                    cat_data["fase"] = "anclas"

    # Calcular totales agregados por fase
    total_excavacion = 0.0
    total_pilas = 0
    total_anclas = 0
    total_muros = 0.0
    tipos_actividades = set()

    for cat_data in categorias.values():
        fase = cat_data["fase"]
        for item in cat_data["items"]:
            unidad_up = item["unidad"].upper()
            cant = item["cantidad"]
            if fase == "excavacion" and unidad_up in ("M3", "M³"):
                total_excavacion += cant
                tipos_actividades.add("excavacion")
            elif fase == "pilas" and unidad_up in ("PZA", "PZAS", "PIEZA", "PIEZAS"):
                total_pilas += int(round(cant))
                tipos_actividades.add("pilas")
            elif fase == "anclas" and unidad_up in ("PZA", "PZAS"):
                total_anclas += int(round(cant))
                tipos_actividades.add("anclas")
            elif fase == "muros" and unidad_up in ("M2", "M²"):
                total_muros += cant
                tipos_actividades.add("muros")

    if not presupuesto_total:
        presupuesto_total = sum(c["importe"] for c in categorias.values())

    # Construir un "frente" único con todas las partidas como actividades planas
    actividades_frente = []
    for cat_data in categorias.values():
        for item in cat_data["items"]:
            actividades_frente.append({
                "descripcion": item["descripcion"],
                "categoria": cat_data["nombre"],
                "cantidad": item["cantidad"],
                "unidad": item["unidad"],
                "importe": item["importe"],
                "tipo": cat_data["fase"],
            })

    frentes = []
    if actividades_frente:
        frentes.append({
            "nombre": nombre_proyecto or "Proyecto",
            "actividades": actividades_frente,
            "tipo_principal": None,
        })

    # Construir desglose de presupuesto compatible con el formato existente
    cat_presupuesto = {}
    mapeo_categoria_presupuesto = {
        "excavacion": "Excavación",
        "pilas": "Cimentación",
        "anclas": "Anclas",
        "muros": "Muros",
        "cimentacion": "Cimentación",
        "generales": "Generales",
        "otros": "Otros",
    }
    for cat_data in categorias.values():
        key = mapeo_categoria_presupuesto.get(cat_data["fase"], "Otros")
        if key not in cat_presupuesto:
            cat_presupuesto[key] = {"total": 0.0, "conceptos": []}
        cat_presupuesto[key]["total"] += cat_data["importe"]
        for item in cat_data["items"]:
            cant_val = item["cantidad"] if item["cantidad"] else None
            imp_val = item["importe"] or 0
            p_unit = (imp_val / cant_val) if (cant_val and cant_val > 0) else None
            cat_presupuesto[key]["conceptos"].append({
                "concepto": item["descripcion"][:200],
                "cantidad": cant_val,
                "unidad": item["unidad"],
                "p_unitario": p_unit,
                "importe": imp_val,
                "progreso": 0,
            })

    num_conceptos_total = sum(len(c["conceptos"]) for c in cat_presupuesto.values())

    # Construir programa semanal agregado por fase
    # Para cada semana, sumar cantidad planeada y importe planeado por fase
    programa_semanal = []
    for s in semanas_cols:
        sem_data = {
            "semana": s["semana"],
            "fecha_inicio": s.get("fecha_inicio"),
            "fecha_fin": s.get("fecha_fin"),
            "excavacion_m3": 0.0,
            "pilas": 0.0,
            "anclas": 0.0,
            "muros_m2": 0.0,
            "presupuesto": 0.0,
            "actividades": [],   # lista de partidas planeadas esta semana
        }
        for cat_data in categorias.values():
            fase = cat_data["fase"]
            for item in cat_data["items"]:
                cant_semana = next(
                    (sw["cantidad"] for sw in item.get("semanas", []) if sw["semana"] == s["semana"]),
                    0.0,
                )
                if cant_semana <= 0:
                    continue
                # Calcular importe proporcional al avance planeado de esta semana
                pct = (cant_semana / item["cantidad"]) if item["cantidad"] else 0
                importe_semana = pct * item["importe"]
                sem_data["presupuesto"] += importe_semana
                sem_data["actividades"].append({
                    "descripcion": item["descripcion"][:140],
                    "categoria": cat_data["nombre"],
                    "fase": fase,
                    "cantidad": cant_semana,
                    "unidad": item["unidad"],
                    "importe": round(importe_semana, 2),
                })

                # Agregar al total por fase
                unidad_up = item["unidad"].upper()
                if fase == "excavacion" and unidad_up in ("M3", "M³"):
                    sem_data["excavacion_m3"] += cant_semana
                elif fase == "pilas" and unidad_up in ("PZA", "PZAS", "PIEZA", "PIEZAS"):
                    sem_data["pilas"] += cant_semana
                elif fase == "anclas" and unidad_up in ("PZA", "PZAS"):
                    sem_data["anclas"] += cant_semana
                elif fase == "muros" and unidad_up in ("M2", "M²"):
                    sem_data["muros_m2"] += cant_semana

        # Redondear
        sem_data["excavacion_m3"] = round(sem_data["excavacion_m3"], 2)
        sem_data["pilas"] = round(sem_data["pilas"], 2)
        sem_data["anclas"] = round(sem_data["anclas"], 2)
        sem_data["muros_m2"] = round(sem_data["muros_m2"], 2)
        sem_data["presupuesto"] = round(sem_data["presupuesto"], 2)
        # Solo incluir semanas con actividad planeada (filtra padding del template Excel)
        tiene_actividad = (
            sem_data["excavacion_m3"] > 0
            or sem_data["pilas"] > 0
            or sem_data["anclas"] > 0
            or sem_data["muros_m2"] > 0
            or sem_data["presupuesto"] > 0
        )
        if tiene_actividad:
            programa_semanal.append(sem_data)

    # Fallback: usar el programa_semanal detectado si no encontramos los metadatos en filas previas
    if not semanas_planeadas and semanas_cols:
        # Usar el total de columnas-semana detectadas (incluye las vacías al final)
        semanas_planeadas = max(s["semana"] for s in semanas_cols)
    if not fecha_inicio and semanas_cols:
        fechas_ini = [s.get("fecha_inicio") for s in semanas_cols if s.get("fecha_inicio")]
        if fechas_ini:
            fecha_inicio = min(fechas_ini)
    if not fecha_fin and semanas_cols:
        fechas_fin = [s.get("fecha_fin") for s in semanas_cols if s.get("fecha_fin")]
        if fechas_fin:
            fecha_fin = max(fechas_fin)

    return {
        "success": True,
        "formato": "programa_obra_v2",
        "nombre_proyecto": nombre_proyecto,
        "frentes": frentes,
        "categorias_detalle": list(categorias.values()),
        "programa_semanal": programa_semanal,
        "presupuesto": {
            "total": round(presupuesto_total, 2),
            "total_general": round(presupuesto_total, 2),
            "num_conceptos": num_conceptos_total,
            "version": "V2",
            "categorias": cat_presupuesto,
            "hoja_seleccionada": ws.title,
            "filename": "programa_obra.xlsx",
        },
        "resumen": {
            "total_frentes": len(frentes),
            "total_actividades": len(actividades_frente),
            "total_dias": semanas_planeadas * 7,
            "semanas_estimadas": semanas_planeadas or 0,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "total_pilas": total_pilas,
            "total_muros": round(total_muros, 2),
            "total_anclas": total_anclas,
            "total_excavacion": round(total_excavacion, 2),
            "tipos_actividades": list(tipos_actividades),
            "semanas_excavacion": 0,
            "semanas_pilas": 0,
            "semanas_muros": 0,
            "semanas_anclas": 0,
            "presupuesto_total": round(presupuesto_total, 2),
            "nombre_proyecto": nombre_proyecto,
        },
    }


def detectar_tipo_actividad(descripcion: str) -> Dict[str, Any]:
    """
    Detecta el tipo de actividad basado en palabras clave en la descripción.
    Retorna el tipo y si aplica anclas.
    """
    desc_lower = descripcion.lower()
    
    # Patrones para detectar tipos
    if any(word in desc_lower for word in ['pila', 'pilas', 'pilote', 'pilotes']):
        return {"tipo": "pilas", "tiene_anclas": True}
    elif any(word in desc_lower for word in ['muro', 'muros', 'muro milan', 'pantalla']):
        return {"tipo": "muros", "tiene_anclas": True}
    elif any(word in desc_lower for word in ['excavac', 'excavar', 'terraceria', 'desmonte', 'despalme']):
        return {"tipo": "excavacion", "tiene_anclas": False}
    elif any(word in desc_lower for word in ['ciment', 'zapata', 'losa', 'dado']):
        return {"tipo": "cimentacion", "tiene_anclas": False}
    elif any(word in desc_lower for word in ['ancla', 'anclas', 'anclaje']):
        return {"tipo": "anclas", "tiene_anclas": True}
    else:
        return {"tipo": "otro", "tiene_anclas": False}


def parse_excel_cronograma(file_content: bytes) -> Dict[str, Any]:
    """
    Parsea un archivo Excel de cronograma y extrae la información de frentes y actividades.
    Detecta automáticamente los tipos de actividades (excavación, pilas, muros, anclas).
    Retorna estructura lista para crear proyecto.

    Soporta dos formatos:
      1. "Programa de Obra V2" (header CONCEPTO|UNIDAD|CANTIDAD|IMPORTE DE CONTRATO,
         con categorías GENERALES/EXCAVACION/ANCLAJE/MUROS/PILAS/etc.).
      2. Formato clásico "FRENTE #PILAS" (legacy).
    """
    # 1) Intentar primero el formato V2 (Avance por Actividades)
    try:
        v2 = parse_excel_programa_obra(file_content)
        if v2 and v2.get("success") and (v2["resumen"]["total_actividades"] > 0):
            return v2
    except Exception as e:
        logging.warning(f"V2 parser falló, intentando legacy: {e}")

    try:
        # Leer Excel
        df = pd.read_excel(BytesIO(file_content), header=None)
        
        frentes = []
        current_frente = None
        # Contadores totales
        total_pilas = 0
        total_muros = 0
        total_anclas = 0
        total_excavacion = 0
        # Fechas
        fecha_inicio_proyecto = None
        fecha_fin_proyecto = None
        total_dias = 0
        # Tipos de actividades detectadas
        tipos_actividades = set()
        # Semanas por tipo
        semanas_por_tipo = {"excavacion": 0, "pilas": 0, "muros": 0, "anclas": 0}
        
        for idx, row in df.iterrows():
            # Detectar encabezado de frente (primera columna contiene "FRENTE")
            first_cell = str(row.iloc[0]).strip().upper() if pd.notna(row.iloc[0]) else ""
            second_cell = str(row.iloc[1]).strip().upper() if len(row) > 1 and pd.notna(row.iloc[1]) else ""
            
            # Si es header de frente
            is_frente_header = first_cell.startswith("FRENTE") and (
                "#PILAS" in second_cell or "PILAS" in second_cell or 
                "#" in second_cell or "CANTIDAD" in second_cell
            )
            
            if is_frente_header:
                # Guardar frente anterior si existe
                if current_frente:
                    frentes.append(current_frente)
                
                current_frente = {
                    "nombre": first_cell,
                    "actividades": [],
                    "tipo_principal": None
                }
                continue
            
            # Si tenemos un frente activo y la segunda columna tiene un número (cantidad)
            if current_frente:
                try:
                    cantidad_val = row.iloc[1] if len(row) > 1 else None
                    if pd.notna(cantidad_val) and isinstance(cantidad_val, (int, float)) and cantidad_val > 0:
                        # Es una actividad válida
                        descripcion = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
                        
                        # Detectar tipo de actividad
                        tipo_info = detectar_tipo_actividad(descripcion)
                        tipo = tipo_info["tipo"]
                        tiene_anclas = tipo_info["tiene_anclas"]
                        
                        # Parsear fechas
                        fecha_inicio_raw = row.iloc[2] if len(row) > 2 else None
                        fecha_fin_raw = row.iloc[3] if len(row) > 3 else None
                        fecha_descabece_raw = row.iloc[4] if len(row) > 4 else None
                        dias_raw = row.iloc[5] if len(row) > 5 else 0
                        
                        # Convertir fechas
                        def parse_fecha(val):
                            if pd.isna(val):
                                return ""
                            if isinstance(val, datetime):
                                return val.strftime("%Y-%m-%d")
                            return excel_date_to_string(val)
                        
                        cantidad = int(cantidad_val)
                        dias = int(dias_raw) if pd.notna(dias_raw) and isinstance(dias_raw, (int, float)) else 0
                        
                        actividad = {
                            "descripcion": descripcion,
                            "cantidad": cantidad,
                            "tipo": tipo,
                            "tiene_anclas": tiene_anclas,
                            "fecha_inicio": parse_fecha(fecha_inicio_raw),
                            "fecha_fin": parse_fecha(fecha_fin_raw),
                            "fecha_descabece": parse_fecha(fecha_descabece_raw),
                            "dias": dias
                        }
                        
                        current_frente["actividades"].append(actividad)
                        
                        # Actualizar contadores según tipo
                        if tipo == "pilas" or "pila" in descripcion.lower():
                            total_pilas += cantidad
                            tipos_actividades.add("pilas")
                            semanas_por_tipo["pilas"] += max(1, dias // 7)
                        elif tipo == "muros":
                            total_muros += cantidad
                            tipos_actividades.add("muros")
                            semanas_por_tipo["muros"] += max(1, dias // 7)
                        elif tipo == "excavacion":
                            total_excavacion += cantidad
                            tipos_actividades.add("excavacion")
                            semanas_por_tipo["excavacion"] += max(1, dias // 7)
                        elif tipo == "anclas":
                            total_anclas += cantidad
                            tipos_actividades.add("anclas")
                            semanas_por_tipo["anclas"] += max(1, dias // 7)
                        else:
                            # Por defecto, si tiene número y parece ser pilas
                            total_pilas += cantidad
                            tipos_actividades.add("pilas")
                        
                        # Si tiene anclas, estimar cantidad
                        if tiene_anclas:
                            total_anclas += cantidad  # Una ancla por pila/muro
                            tipos_actividades.add("anclas")
                        
                        total_dias += dias
                        
                        # Actualizar fechas del proyecto
                        if actividad["fecha_inicio"]:
                            if not fecha_inicio_proyecto or actividad["fecha_inicio"] < fecha_inicio_proyecto:
                                fecha_inicio_proyecto = actividad["fecha_inicio"]
                        
                        if actividad["fecha_descabece"]:
                            if not fecha_fin_proyecto or actividad["fecha_descabece"] > fecha_fin_proyecto:
                                fecha_fin_proyecto = actividad["fecha_descabece"]
                        elif actividad["fecha_fin"]:
                            if not fecha_fin_proyecto or actividad["fecha_fin"] > fecha_fin_proyecto:
                                fecha_fin_proyecto = actividad["fecha_fin"]
                                
                except Exception as e:
                    logging.warning(f"Error parseando fila {idx}: {e}")
                    continue
        
        # Agregar último frente
        if current_frente:
            frentes.append(current_frente)
        
        # Calcular semanas basado en el rango de fechas
        semanas_estimadas = 1
        if fecha_inicio_proyecto and fecha_fin_proyecto:
            try:
                inicio = datetime.strptime(fecha_inicio_proyecto, "%Y-%m-%d")
                fin = datetime.strptime(fecha_fin_proyecto, "%Y-%m-%d")
                dias_totales = (fin - inicio).days
                semanas_estimadas = max(1, (dias_totales + 6) // 7)  # Redondear hacia arriba
            except Exception:
                semanas_estimadas = max(1, total_dias // 7) if total_dias > 0 else len(frentes) * 4
        
        return {
            "success": True,
            "frentes": frentes,
            "resumen": {
                "total_frentes": len(frentes),
                "total_actividades": sum(len(f["actividades"]) for f in frentes),
                "total_dias": total_dias,
                "semanas_estimadas": semanas_estimadas,
                "fecha_inicio": fecha_inicio_proyecto,
                "fecha_fin": fecha_fin_proyecto,
                # Métricas por tipo
                "total_pilas": total_pilas,
                "total_muros": total_muros,
                "total_anclas": total_anclas,
                "total_excavacion": total_excavacion,
                # Tipos de actividades detectadas
                "tipos_actividades": list(tipos_actividades),
                # Semanas por tipo
                "semanas_excavacion": semanas_por_tipo["excavacion"],
                "semanas_pilas": semanas_por_tipo["pilas"],
                "semanas_muros": semanas_por_tipo["muros"],
                "semanas_anclas": semanas_por_tipo["anclas"]
            }
        }
        
    except Exception as e:
        logging.error(f"Error parseando Excel: {e}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": str(e)
        }


async def analizar_foto_avance(
    imagen_base64: str,
    imagen_anterior_base64: Optional[str] = None,
    pilas_planeadas: int = 0,
    anclas_planeadas: int = 0,
    semana_actual: int = 1
) -> Dict[str, Any]:
    """
    Analiza una foto de avance de obra usando Gemini Vision para detectar:
    - Número de pilas visibles
    - Número de anclas instaladas
    - Comparación con semana anterior
    - Pronóstico de avance
    """
    import tempfile
    import os
    
    try:
        from emergentintegrations.llm.chat import LlmChat, FileContentWithMimeType
        
        if not EMERGENT_LLM_KEY:
            return {"success": False, "error": "EMERGENT_LLM_KEY no configurada"}
        
        # Guardar imagen temporalmente
        import base64
        image_data = base64.b64decode(imagen_base64)
        
        with tempfile.NamedTemporaryFile(suffix='.jpg', delete=False) as tmp_file:
            tmp_file.write(image_data)
            temp_path = tmp_file.name
        
        try:
            # Construir prompt detallado
            prompt = f"""Analiza esta imagen aérea de una obra de construcción de cimentaciones profundas.

CONTEXTO:
- Semana actual del proyecto: {semana_actual}
- Pilas planeadas hasta esta semana: {pilas_planeadas}
- Anclas planeadas hasta esta semana: {anclas_planeadas}

QUÉ BUSCAR:
1. PILAS DE CIMENTACIÓN: Son columnas circulares de concreto que emergen del suelo, típicamente de 60cm a 120cm de diámetro. Se ven como círculos o cilindros grises desde arriba.

2. ANCLAS/ANCLAJES: Son barras de acero (varillas) que sobresalen de las pilas ya construidas. Se ven como puntos oscuros o pequeñas estructuras metálicas encima de las pilas terminadas.

3. PILAS EN PROCESO: Áreas donde se está excavando (agujeros en el suelo) o donde hay maquinaria de perforación trabajando.

4. EXCAVACIÓN: Áreas donde se ha removido tierra, taludes, rampas de acceso.

INSTRUCCIONES:
- Cuenta TODOS los elementos que puedas identificar claramente
- Si hay duda, incluye el elemento pero indica confianza media/baja
- Describe qué ves en la imagen

Responde EXACTAMENTE en este formato JSON:
{{
    "pilas_detectadas": <número de pilas terminadas visibles>,
    "anclas_detectadas": <número de anclas/anclajes visibles>,
    "pilas_en_proceso": <número de pilas en construcción>,
    "excavaciones_activas": <número de puntos de excavación>,
    "porcentaje_avance_estimado": <porcentaje 0-100 basado en lo visible>,
    "estado_proyecto": "<EN_TIEMPO | ADELANTADO | RETRASADO | NO_DETERMINABLE>",
    "confianza_deteccion": "<ALTA | MEDIA | BAJA>",
    "elementos_identificados": "<lista de lo que se puede ver claramente>",
    "observaciones": "<descripción detallada de lo que se observa en la imagen>",
    "condiciones_terreno": "<estado del terreno, clima visible, etc>",
    "maquinaria_visible": "<descripción de equipos visibles>",
    "recomendaciones": "<sugerencias basadas en lo observado>"
}}"""
            
            # Crear chat con Gemini Vision
            llm = LlmChat(
                api_key=EMERGENT_LLM_KEY,
                session_id=f"analisis-foto-{datetime.now().strftime('%Y%m%d%H%M%S')}",
                system_message="Eres un experto en análisis de imágenes de construcción civil."
            ).with_model("gemini", "gemini-2.0-flash")
            
            # Preparar imagen como FileContentWithMimeType usando la ruta del archivo temporal
            image_file = FileContentWithMimeType(
                mime_type="image/jpeg",
                file_path=temp_path
            )
            
            # Crear mensaje con texto e imagen
            from emergentintegrations.llm.chat import UserMessage
            user_message = UserMessage(
                text=prompt,
                file_contents=[image_file]
            )
            
            # Enviar mensaje
            response = await llm.send_message(user_message)
            
            # Parsear respuesta JSON
            import json
            import re
            
            # Limpiar respuesta
            response_text = response.strip()
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            if response_text.startswith("```"):
                response_text = response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            
            # Extraer JSON de la respuesta
            json_match = re.search(r'\{[\s\S]*\}', response_text)
            if json_match:
                result = json.loads(json_match.group())
                result["success"] = True
                result["raw_response"] = response
                return result
            else:
                return {
                    "success": True,
                    "raw_response": response,
                    "pilas_detectadas": 0,
                    "anclas_detectadas": 0,
                    "observaciones": response,
                    "confianza_deteccion": "BAJA"
                }
        finally:
            # Limpiar archivo temporal
            if os.path.exists(temp_path):
                os.unlink(temp_path)
            
    except Exception as e:
        logging.error(f"Error analizando imagen: {e}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": str(e)
        }


async def generar_reporte_progreso(
    proyecto_nombre: str,
    semana_actual: int,
    pilas_planeadas: int,
    pilas_detectadas: int,
    anclas_planeadas: int,
    anclas_detectadas: int,
    historial_semanas: List[Dict] = None
) -> Dict[str, Any]:
    """
    Genera un reporte de progreso usando IA
    """
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        if not EMERGENT_LLM_KEY:
            return {"success": False, "error": "EMERGENT_LLM_KEY no configurada"}
        
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"reporte-{datetime.now().isoformat()}",
            system_message="Eres un experto en gestión de proyectos de construcción."
        ).with_model("gemini", "gemini-2.5-flash")
        
        historial_texto = ""
        if historial_semanas:
            historial_texto = "\nHistorial de semanas anteriores:\n"
            for h in historial_semanas:
                historial_texto += f"- Semana {h.get('semana')}: {h.get('pilas_detectadas', 0)} pilas, {h.get('anclas_detectadas', 0)} anclas\n"
        
        prompt = f"""Genera un breve reporte de progreso para el proyecto "{proyecto_nombre}":

Semana actual: {semana_actual}
Pilas planeadas: {pilas_planeadas}
Pilas detectadas: {pilas_detectadas}
Anclas planeadas: {anclas_planeadas}
Anclas detectadas: {anclas_detectadas}
{historial_texto}

Responde en JSON con:
{{
    "resumen_ejecutivo": "<2-3 oraciones>",
    "porcentaje_pilas": <número>,
    "porcentaje_anclas": <número>,
    "estado": "<EN_TIEMPO | ADELANTADO | RETRASADO>",
    "dias_diferencia": <positivo si adelantado, negativo si retrasado>,
    "fecha_estimada_termino": "<si hay retraso, nueva fecha estimada>",
    "acciones_recomendadas": ["<acción 1>", "<acción 2>"]
}}"""
        
        response = await chat.send_message(UserMessage(text=prompt))
        
        import json
        import re
        json_match = re.search(r'\{[\s\S]*\}', response)
        if json_match:
            result = json.loads(json_match.group())
            result["success"] = True
            return result
        
        return {"success": True, "resumen_ejecutivo": response}
        
    except Exception as e:
        logging.error(f"Error generando reporte: {e}")
        return {"success": False, "error": str(e)}
