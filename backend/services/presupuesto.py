"""
Presupuesto Service - DrON Topografía

Extrae presupuestos desde archivos Excel (.xlsx/.xlsm) usando IA Gemini para
clasificar conceptos en categorías estándar (Generales, Excavación, Cimentación,
Anclas, Muros, Edificación, Otros).

Detecta automáticamente hojas múltiples (R3/R4/PPTO/etc.) y permite al usuario
elegir cuál usar.
"""
import io
import json
import logging
import re
import uuid
from typing import Optional

from openpyxl import load_workbook
from emergentintegrations.llm.chat import LlmChat, UserMessage

from core.config import EMERGENT_LLM_KEY

logger = logging.getLogger(__name__)

# Categorías estándar usadas en proyectos DrON
CATEGORIAS = ["Generales", "Excavación", "Cimentación", "Anclas", "Muros", "Edificación", "Otros"]


def listar_hojas(excel_bytes: bytes) -> list:
    """Lee el Excel y devuelve la lista de hojas con metadatos básicos."""
    try:
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True, keep_vba=False)
        hojas = []
        for name in wb.sheetnames:
            ws = wb[name]
            # Detectar si parece presupuesto (busca palabras clave en primeras 25 filas
            # para capturar APU donde los headers aparecen en R13+)
            sample_text = ""
            for row in ws.iter_rows(min_row=1, max_row=25, values_only=True):
                for c in row:
                    if c is not None:
                        sample_text += str(c) + " "
            sample_lower = sample_text.lower()
            es_presupuesto = any(
                kw in sample_lower for kw in [
                    "concepto", "importe", "p. unitario", "p.u.", "subtotal",
                    "total", "presupuesto", "cantidad", "unidad",
                    # APU markers
                    "partida:", "análisis:", "analisis:", "precios unitarios",
                ]
            )
            # Heurística para identificar versiones (R3, R4, PPTO, Neodata, etc.)
            posible_version = None
            m = re.search(r"\bR\s*\d+\b", name, re.IGNORECASE)
            if m:
                posible_version = m.group(0).upper().replace(" ", "")
            elif "ppto" in name.lower() or "presupuesto" in name.lower():
                posible_version = "PPTO"
            hojas.append({
                "nombre": name,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "es_presupuesto": es_presupuesto,
                "posible_version": posible_version,
            })
        return hojas
    except Exception as e:
        logger.exception("Error listando hojas")
        raise ValueError(f"No se pudo leer el Excel: {e}")


def extraer_filas_planas(excel_bytes: bytes, sheet_name: str) -> list:
    """Convierte una hoja en una lista de filas no vacías, lista para enviar a la IA."""
    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True, keep_vba=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Hoja '{sheet_name}' no existe. Disponibles: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        clean = [str(c).strip() if c is not None else "" for c in row]
        if any(clean):
            rows.append(clean)
    return rows


def detectar_columnas_relevantes(rows: list, version: Optional[str] = None) -> dict:
    """
    Detecta las columnas: Concepto, Unidad, Cantidad, P.Unitario, Importe.
    En hojas tipo COMPARATIVA puede haber varias columnas Cantidad/Importe (R3 vs R4).
    """
    # Buscar header row (donde aparezcan "concepto", "unidad", etc.)
    header_idx = None
    for i, row in enumerate(rows[:25]):
        joined = " ".join(c.lower() for c in row)
        if "concepto" in joined and ("unidad" in joined or "cantidad" in joined):
            header_idx = i
            break
    if header_idx is None:
        return {"header_idx": None}

    header = rows[header_idx]
    cols = {
        "concepto": None,
        "unidad": None,
        "cantidad": None,
        "p_unitario": None,
        "importe": None,
        "header_idx": header_idx,
    }
    # Si versión especificada, intentar localizar columnas específicas R3/R4
    version_target = (version or "").upper().strip()
    
    for idx, cell in enumerate(header):
        cell_l = (cell or "").lower().strip()
        if "concepto" in cell_l and cols["concepto"] is None:
            cols["concepto"] = idx
        elif "unidad" in cell_l and cols["unidad"] is None:
            cols["unidad"] = idx
        elif ("cantidad" in cell_l or "cant." in cell_l) and cols["cantidad"] is None:
            cols["cantidad"] = idx
        elif ("p. unitario" in cell_l or "p.u." in cell_l or "precio unitario" in cell_l) and cols["p_unitario"] is None:
            cols["p_unitario"] = idx
        elif "importe" in cell_l and cols["importe"] is None:
            cols["importe"] = idx
    
    # Si comparativa: buscar marcadores de versión en la fila anterior al header
    if version_target and header_idx > 0:
        marker_row = rows[header_idx - 1]
        # Encuentra qué columna inicia la versión target
        for idx, cell in enumerate(marker_row):
            if cell and version_target in cell.upper().replace(" ", ""):
                # Re-asignar cantidad e importe a las columnas a la derecha del marker
                for j in range(idx, len(header)):
                    hl = (header[j] or "").lower()
                    if ("cantidad" in hl or "cant." in hl) and j >= idx:
                        cols["cantidad"] = j
                        break
                for j in range(idx, len(header)):
                    hl = (header[j] or "").lower()
                    if "importe" in hl and j >= idx:
                        cols["importe"] = j
                        break
                break
    
    return cols


def _normalizar_categoria_apu(nombre: str) -> str:
    """Normaliza el nombre de la partida APU a una categoría estándar DrON."""
    n = (nombre or "").upper().strip()
    if any(k in n for k in ["EXCAVAC", "RELLENO", "CARGA", "RETIRO"]):
        return "Excavación"
    if any(k in n for k in ["ANCLA", "ANCLAJE", "ESTABILIZAC", "TORON"]):
        return "Anclas"
    if any(k in n for k in ["MURO", "RECUBRIMIENTO", "LANZADO"]):
        return "Muros"
    if any(k in n for k in ["PILA", "REFORZAMIENTO", "COLINDANCIA", "PERFORAC", "PERFIL"]):
        return "Cimentación"
    if any(k in n for k in ["ZAPATA", "CIMENT", "LOSA", "TRABE", "DALA"]):
        return "Cimentación"
    if any(k in n for k in ["TOPOGRAF", "TRASLAD", "MOVILIZAC", "MANIOBRA", "FLETE"]):
        return "Generales"
    if any(k in n for k in ["EDIFICAC", "ESTRUCT", "COLUMNA", "ENTREPISO"]):
        return "Edificación"
    return "Otros"


def parse_excel_apu(excel_bytes: bytes) -> Optional[dict]:
    """
    Parser para archivos "Análisis de Precios Unitarios" (APU/PU).

    Estructura típica:
      • Encabezado del análisis: columnas Código | Concepto | Unidad | P. Unitario | Op. | Cantidad | Importe | %
      • Filas de partida:
          Partida:   <NOMBRE CATEGORÍA>    Análisis No.:  <n>
          Análisis:  <CÓDIGO>  <unidad>  <cantidad>  <importe>
          <descripción>
          ... bloques BÁSICOS / MATERIALES / MANO DE OBRA / EQUIPO Y HERRAMIENTA ...
          SUBTOTAL: ...
          (CD) Costo directo
          PRECIO UNITARIO
          (* texto *)

    Devuelve un dict con la misma estructura que extraer_presupuesto_con_ia(),
    o None si la hoja no parece formato APU.
    """
    try:
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True, keep_vba=False)
    except Exception:
        return None

    # Buscar la primera hoja con marcadores "Partida:" + "Análisis:"
    ws = None
    for name in wb.sheetnames:
        candidate = wb[name]
        encontrado = False
        for r in range(1, min(60, candidate.max_row + 1)):
            for c in range(1, 4):
                v = candidate.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip().upper().startswith("PARTIDA:"):
                    encontrado = True
                    break
            if encontrado:
                break
        if encontrado:
            ws = candidate
            break
    if not ws:
        return None

    # Detectar columnas a partir del header "Código | Concepto | Unidad | P. Unitario | Op. | Cantidad | Importe"
    header_row = None
    for r in range(1, min(40, ws.max_row + 1)):
        vals = [str(ws.cell(row=r, column=c).value or "").upper() for c in range(1, 9)]
        joined = " | ".join(vals)
        if "CÓDIGO" in joined and "CONCEPTO" in joined and "UNIDAD" in joined and "IMPORTE" in joined:
            header_row = r
            break
    if not header_row:
        return None

    col_concepto = col_unidad = col_pu = col_cantidad = col_importe = None
    for c in range(1, 10):
        h = str(ws.cell(row=header_row, column=c).value or "").upper().strip()
        if "CONCEPTO" in h:
            col_concepto = c
        elif "UNIDAD" in h:
            col_unidad = c
        elif "P. UNITARIO" in h or "PRECIO UNITARIO" in h:
            col_pu = c
        elif "CANTIDAD" in h:
            col_cantidad = c
        elif "IMPORTE" in h:
            col_importe = c

    if not (col_concepto and col_importe):
        return None

    # Recorrer y extraer cada análisis
    categorias = {}
    categoria_actual = None

    for r in range(header_row + 1, ws.max_row + 1):
        v_a = ws.cell(row=r, column=1).value
        v_b = ws.cell(row=r, column=col_concepto).value
        v_a_str = str(v_a or "").strip().upper()
        v_b_str = str(v_b or "").strip()

        # 1) Cabecera de partida: A="Partida:" → B = nombre categoría
        if v_a_str.startswith("PARTIDA"):
            categoria_actual = v_b_str
            continue

        # 2) Cabecera de análisis: A="Análisis:" → B = código, C=unidad, D=cantidad, E=importe (col_pu en este caso es D)
        if v_a_str.startswith("ANÁLISIS") or v_a_str.startswith("ANALISIS"):
            # Columnas en esta variante: B=código, D=unidad, F=cantidad, G=importe
            codigo = v_b_str
            unidad = str(ws.cell(row=r, column=col_unidad).value or "").strip() if col_unidad else ""
            # En el header de análisis, "P. Unitario" suele tener la unidad (M3)
            if not unidad or len(unidad) > 6:
                # caer al col_pu si el campo unidad no es legible
                alt_unidad = str(ws.cell(row=r, column=col_pu).value or "").strip() if col_pu else ""
                if alt_unidad and len(alt_unidad) <= 6:
                    unidad = alt_unidad
            cantidad = ws.cell(row=r, column=col_cantidad).value if col_cantidad else None
            importe = ws.cell(row=r, column=col_importe).value if col_importe else None

            try:
                cant_f = float(cantidad) if isinstance(cantidad, (int, float)) else None
            except Exception:
                cant_f = None
            try:
                imp_f = float(importe) if isinstance(importe, (int, float)) else None
            except Exception:
                imp_f = None

            if not categoria_actual or imp_f is None or imp_f == 0:
                continue

            # Recopilar descripción de las siguientes 1-3 filas (texto largo en columna B)
            descripcion = ""
            for dr in range(r + 1, min(r + 4, ws.max_row + 1)):
                desc_val = ws.cell(row=dr, column=1).value
                if isinstance(desc_val, str) and len(desc_val) > 30:
                    descripcion = desc_val.strip()
                    break

            cat_estandar = _normalizar_categoria_apu(categoria_actual)
            if cat_estandar not in categorias:
                categorias[cat_estandar] = {"total": 0.0, "conceptos": []}

            p_unit = (imp_f / cant_f) if cant_f else None
            categorias[cat_estandar]["conceptos"].append({
                "concepto": (descripcion or codigo)[:300],
                "unidad": unidad,
                "cantidad": cant_f,
                "p_unitario": round(p_unit, 2) if p_unit is not None else None,
                "importe": round(imp_f, 2),
                "seccion_excel": categoria_actual,
                "codigo": codigo,
            })
            categorias[cat_estandar]["total"] += imp_f
            continue

    if not categorias:
        return None

    # Redondear totales
    for cat in categorias.values():
        cat["total"] = round(cat["total"], 2)
    total_general = round(sum(c["total"] for c in categorias.values()), 2)
    num_conceptos = sum(len(c["conceptos"]) for c in categorias.values())

    return {
        "version": "APU",
        "sheet": ws.title,
        "categorias": categorias,
        "total_general": total_general,
        "num_conceptos": num_conceptos,
        "formato": "apu",
    }


async def extraer_presupuesto_con_ia(
    excel_bytes: bytes,
    sheet_name: str,
    version: Optional[str] = None,
) -> dict:
    """
    Extrae el presupuesto completo desde una hoja específica usando IA Gemini
    para clasificar cada concepto en una categoría estándar.

    Returns:
        {
            "version": "R4",
            "sheet": "R4",
            "categorias": {
                "Excavación": {
                    "total": 2060416.92,
                    "conceptos": [
                        {"concepto": "Excavación por medios mecánicos",
                         "unidad": "M3", "cantidad": 12340.7,
                         "p_unitario": 63.69, "importe": 785979.18},
                        ...
                    ]
                },
                "Cimentación": {...},
                ...
            },
            "total_general": 5740000.0,
        }
    """
    # 1) Intentar primero el parser APU (Análisis de Precios Unitarios), que es
    # determinista y no requiere llamada a la IA. Si funciona, lo usamos.
    try:
        apu = parse_excel_apu(excel_bytes)
        if apu and apu["num_conceptos"] > 0:
            logger.info(f"Formato APU detectado: {apu['num_conceptos']} conceptos, total ${apu['total_general']:,.2f}")
            return apu
    except Exception as e:
        logger.warning(f"Parser APU falló, intentando flujo IA: {e}")

    rows = extraer_filas_planas(excel_bytes, sheet_name)
    if not rows:
        raise ValueError(f"La hoja '{sheet_name}' está vacía")

    cols = detectar_columnas_relevantes(rows, version=version)
    if cols.get("header_idx") is None or cols.get("concepto") is None:
        raise ValueError(
            "No se pudo detectar la estructura del presupuesto. "
            "Asegúrate de que la hoja tiene encabezados como 'Concepto', 'Unidad', 'Cantidad', 'P. Unitario', 'Importe'."
        )

    # Extraer renglones de conceptos (skip header y filas de subtotales si son obvias)
    conceptos = []
    seccion_actual = None
    header_idx = cols["header_idx"]
    for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        concepto_txt = row[cols["concepto"]] if cols["concepto"] < len(row) else ""
        unidad_txt = row[cols["unidad"]] if cols["unidad"] is not None and cols["unidad"] < len(row) else ""
        cant_raw = row[cols["cantidad"]] if cols["cantidad"] is not None and cols["cantidad"] < len(row) else ""
        pu_raw = row[cols["p_unitario"]] if cols["p_unitario"] is not None and cols["p_unitario"] < len(row) else ""
        imp_raw = row[cols["importe"]] if cols["importe"] is not None and cols["importe"] < len(row) else ""

        if not concepto_txt.strip():
            continue
        # Detectar headers de sección (sin unidad ni cantidad, mayúsculas)
        if (not unidad_txt and not cant_raw and concepto_txt.strip().isupper()
                and len(concepto_txt) < 40):
            seccion_actual = concepto_txt.strip()
            continue
        # Detectar subtotales/totales (saltarlos)
        cl = concepto_txt.lower()
        if any(kw in cl for kw in ["subtotal", "total ", "iva", "suma"]) and len(concepto_txt) < 40:
            continue

        # Parsear números
        def to_float(v):
            try:
                return float(str(v).replace(",", "").replace("$", "").strip())
            except Exception:
                return None

        cant = to_float(cant_raw)
        pu = to_float(pu_raw)
        imp = to_float(imp_raw)
        if imp is None and cant is not None and pu is not None:
            imp = cant * pu
        if imp is None or imp == 0:
            continue

        conceptos.append({
            "row": i,
            "concepto": concepto_txt.strip(),
            "unidad": unidad_txt.strip(),
            "cantidad": cant,
            "p_unitario": pu,
            "importe": round(imp, 2),
            "seccion_excel": seccion_actual,
        })

    if not conceptos:
        raise ValueError("No se encontraron conceptos válidos en la hoja")

    logger.info(f"Extraídos {len(conceptos)} conceptos. Clasificando con IA…")

    # ---- IA: clasificar cada concepto en categoría estándar ----
    # Empacar conceptos en JSON compacto para enviar al LLM
    payload = [
        {
            "id": i,
            "concepto": c["concepto"][:120],
            "unidad": c["unidad"],
            "seccion": c["seccion_excel"] or "",
        }
        for i, c in enumerate(conceptos)
    ]

    prompt = f"""Tengo {len(conceptos)} conceptos de un presupuesto de obra de construcción mexicana.
Tu tarea: clasificar CADA concepto en UNA de estas categorías estándar (responde con la categoría exacta):

- Generales: traslado, topografía, maniobras, fletes, equipo, herramientas.
- Excavación: excavación, carga, retiro de material, terracerías.
- Cimentación: pilas, perforación, lanzado de concreto, reforzamiento estructural.
- Anclas: anclas, fijaciones de ancla, ferretería.
- Muros: muros, lanzado, recubrimiento de muro, malla de muro.
- Edificación: estructura, losas, columnas, edificación general.
- Otros: cualquiera que no encaje claramente.

Devuelve **únicamente** un JSON array con este formato exacto:
[{{"id": 0, "categoria": "Excavación"}}, {{"id": 1, "categoria": "Cimentación"}}, ...]

NO incluyas explicaciones, texto adicional ni markdown. Solo el JSON.

Conceptos:
{json.dumps(payload, ensure_ascii=False)}"""

    try:
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"presupuesto-{uuid.uuid4().hex[:8]}",
            system_message="Eres experto en costos y presupuestos de obra civil en México. Clasificas conceptos en categorías con precisión."
        ).with_model("gemini", "gemini-2.5-flash")
        response = await chat.send_message(UserMessage(text=prompt))
        text = response.strip()
        # Limpiar markdown si lo trae
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        clasificaciones = json.loads(text)
    except Exception:
        logger.exception("Error parseando respuesta IA, usando fallback heurístico")
        clasificaciones = []
        for i, c in enumerate(conceptos):
            cat = _clasificar_heuristico(c["concepto"], c["seccion_excel"])
            clasificaciones.append({"id": i, "categoria": cat})

    # Mapear de id -> categoría
    map_cat = {cl["id"]: cl["categoria"] for cl in clasificaciones if "id" in cl}

    # Agrupar por categoría
    categorias = {cat: {"total": 0.0, "conceptos": []} for cat in CATEGORIAS}
    for i, c in enumerate(conceptos):
        cat = map_cat.get(i, _clasificar_heuristico(c["concepto"], c["seccion_excel"]))
        if cat not in categorias:
            cat = "Otros"
        c_clean = {k: v for k, v in c.items() if k != "row"}
        categorias[cat]["conceptos"].append(c_clean)
        categorias[cat]["total"] += c["importe"]
    # Quitar categorías vacías
    categorias = {k: v for k, v in categorias.items() if v["conceptos"]}
    for cat in categorias.values():
        cat["total"] = round(cat["total"], 2)

    total_general = round(sum(c["total"] for c in categorias.values()), 2)

    return {
        "version": version or sheet_name,
        "sheet": sheet_name,
        "categorias": categorias,
        "total_general": total_general,
        "num_conceptos": len(conceptos),
    }


def _clasificar_heuristico(concepto: str, seccion: Optional[str]) -> str:
    """Fallback simple sin IA: clasificación por palabras clave."""
    c = (concepto + " " + (seccion or "")).lower()
    if any(k in c for k in ["traslado", "topograf", "manibra", "maniobra", "flete", "equipo"]):
        return "Generales"
    if any(k in c for k in ["excavac", "carga y retiro", "material suelto", "terrac"]):
        return "Excavación"
    if any(k in c for k in ["pila", "perforac", "reforzam", "concreto lanz"]):
        return "Cimentación"
    if "ancla" in c:
        return "Anclas"
    if "muro" in c:
        return "Muros"
    if any(k in c for k in ["losa", "columna", "edificac", "estructura"]):
        return "Edificación"
    return "Otros"
